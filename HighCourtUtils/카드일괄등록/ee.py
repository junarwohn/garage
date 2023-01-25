import openpyxl
FILE_PATH = input()
if FILE_PATH[0] == '"' and FILE_PATH[-1] == '"':
    FILE_PATH = FILE_PATH[1:-1]

li = []
cheif_list = []
cheif_list_name = []
adt_list = []
wb = openpyxl.load_workbook(FILE_PATH)
sheet_list = wb.sheetnames[1:]
for sl in sheet_list:
    ws = wb[sl]
    for r in ws.rows:
    
     #   line = str(r[0].value) + " " + str(r[2].value) + " " + str(r[1].value) + "\n"
        element = [str(r[0].value), str(r[2].value), str(r[1].value)]
        cheif_list.append(element)
      # for i in range(len(r)):
      #     result_str += "'" + str(r[i].value) + "',"
      # result_str = result_str[:-1]
      # result_str += ")\n"

     #   li.append(line)

FILE_PATH = input()
wb = openpyxl.load_workbook(FILE_PATH)
ws = wb['Sheet1']
for r in ws.rows:
    element = [str(r[1].value), str(r[0].value)]
    adt_list.append(element)
  # for i in range(len(r)):
  #     result_str += "'" + str(r[i].value) + "',"
  # result_str = result_str[:-1]
  # result_str += ")\n"

   # li.append(line)

for adt_e in adt_list:
    is_found = False
    for c_e in cheif_list:
        if adt_e[0] in c_e:
            for valid in adt_list:
                if c_e[0] in valid:
                    print(c_e)
                    is_found = True
    if not is_found:
        print(["", adt_e[0], ""])
