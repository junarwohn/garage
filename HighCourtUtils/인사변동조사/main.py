# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'main.ui'
#
# Created by: PyQt5 UI code generator 5.9.1
#
# WARNING! All changes made in this file will be lost!

from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl
import csv
from dialog_license import Ui_Dialog as DialogLicense
from doomsLi import Ui_Dialog as InfoLicense
from charmGongik import Ui_Dialog as CharmGongik
from exceptionDialog import Ui_Dialog as ExceptionDialog

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(997, 860)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.centralwidget)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.tableWidget_main = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget_main.setObjectName("tableWidget_main")
        self.tableWidget_main.setColumnCount(0)
        self.tableWidget_main.setRowCount(0)
        self.verticalLayout_3.addWidget(self.tableWidget_main)
        self.New_Main_File_btn = QtWidgets.QPushButton(self.centralwidget)
        self.New_Main_File_btn.setObjectName("New_Main_File_btn")
        self.verticalLayout_3.addWidget(self.New_Main_File_btn)
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_6.setObjectName("horizontalLayout_6")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setObjectName("label")
        self.horizontalLayout_6.addWidget(self.label)
        self.spinBox_Main_name_pos = QtWidgets.QSpinBox(self.centralwidget)
        self.spinBox_Main_name_pos.setObjectName("spinBox_Main_name_pos")
        self.horizontalLayout_6.addWidget(self.spinBox_Main_name_pos)
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout_6.addWidget(self.label_3)
        self.spinBox_Main_birth_pos = QtWidgets.QSpinBox(self.centralwidget)
        self.spinBox_Main_birth_pos.setObjectName("spinBox_Main_birth_pos")
        self.horizontalLayout_6.addWidget(self.spinBox_Main_birth_pos)
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setObjectName("label_5")
        self.horizontalLayout_6.addWidget(self.label_5)
        self.spinBox_Main_target_pos = QtWidgets.QSpinBox(self.centralwidget)
        self.spinBox_Main_target_pos.setObjectName("spinBox_Main_target_pos")
        self.horizontalLayout_6.addWidget(self.spinBox_Main_target_pos)
        self.verticalLayout_3.addLayout(self.horizontalLayout_6)
        self.horizontalLayout_2.addLayout(self.verticalLayout_3)
        self.verticalLayout_4 = QtWidgets.QVBoxLayout()
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.tableWidget_sub = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget_sub.setObjectName("tableWidget_sub")
        self.tableWidget_sub.setColumnCount(0)
        self.tableWidget_sub.setRowCount(0)
        self.verticalLayout_4.addWidget(self.tableWidget_sub)
        self.New_Sub_File_btn = QtWidgets.QPushButton(self.centralwidget)
        self.New_Sub_File_btn.setObjectName("New_Sub_File_btn")
        self.verticalLayout_4.addWidget(self.New_Sub_File_btn)
        self.horizontalLayout_7 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_7.setObjectName("horizontalLayout_7")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_7.addWidget(self.label_2)
        self.spinBox_Sub_name_pos = QtWidgets.QSpinBox(self.centralwidget)
        self.spinBox_Sub_name_pos.setObjectName("spinBox_Sub_name_pos")
        self.horizontalLayout_7.addWidget(self.spinBox_Sub_name_pos)
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setObjectName("label_4")
        self.horizontalLayout_7.addWidget(self.label_4)
        self.spinBox_Sub_birth_pos = QtWidgets.QSpinBox(self.centralwidget)
        self.spinBox_Sub_birth_pos.setObjectName("spinBox_Sub_birth_pos")
        self.horizontalLayout_7.addWidget(self.spinBox_Sub_birth_pos)
        self.verticalLayout_4.addLayout(self.horizontalLayout_7)
        self.horizontalLayout_2.addLayout(self.verticalLayout_4)
        self.verticalLayout.addLayout(self.horizontalLayout_2)
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.Apply_btn = QtWidgets.QPushButton(self.centralwidget)
        self.Apply_btn.setObjectName("Apply_btn")
        self.horizontalLayout_5.addWidget(self.Apply_btn)
        self.Save_btn = QtWidgets.QPushButton(self.centralwidget)
        self.Save_btn.setObjectName("Save_btn")
        self.horizontalLayout_5.addWidget(self.Save_btn)
        self.verticalLayout_2.addLayout(self.horizontalLayout_5)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.tableWidget_Apply = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget_Apply.setObjectName("tableWidget_Apply")
        self.tableWidget_Apply.setColumnCount(0)
        self.tableWidget_Apply.setRowCount(0)
        self.horizontalLayout_3.addWidget(self.tableWidget_Apply)
        self.verticalLayout_2.addLayout(self.horizontalLayout_3)
        self.verticalLayout.addLayout(self.verticalLayout_2)
        self.CharmGongik_btn = QtWidgets.QPushButton(self.centralwidget)
        self.CharmGongik_btn.setObjectName("CharmGongik_btn")
        self.verticalLayout.addWidget(self.CharmGongik_btn)
        self.horizontalLayout.addLayout(self.verticalLayout)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 997, 21))
        self.menubar.setObjectName("menubar")
        self.menuInfo = QtWidgets.QMenu(self.menubar)
        self.menuInfo.setObjectName("menuInfo")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionQT_License = QtWidgets.QAction(MainWindow)
        self.actionQT_License.setObjectName("actionQT_License")
        self.actionDeveloper_License = QtWidgets.QAction(MainWindow)
        self.actionDeveloper_License.setObjectName("actionDeveloper_License")
        self.menuInfo.addAction(self.actionQT_License)
        self.menuInfo.addAction(self.actionDeveloper_License)
        self.menubar.addAction(self.menuInfo.menuAction())

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.CharmGongik_btn.clicked.connect(self.ChamGongikView)
        self.actionQT_License.triggered.connect(self.openSourceLicenseView)
        self.actionDeveloper_License.triggered.connect(self.doomsheartLicenseView)
        self.New_Main_File_btn.clicked.connect(self.openNewMainFileDialog)
        self.New_Sub_File_btn.clicked.connect(self.openNewSubFileDialog)
        self.Apply_btn.clicked.connect(self.applyResult)
        self.Save_btn.clicked.connect(self.saveResult)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.New_Main_File_btn.setText(_translate("MainWindow", "New Main File"))
        self.label.setText(_translate("MainWindow", "NamePosition"))
        self.label_3.setText(_translate("MainWindow", "BirthPosition"))
        self.label_5.setText(_translate("MainWindow", "TargetPosition"))
        self.New_Sub_File_btn.setText(_translate("MainWindow", "New Sub File"))
        self.label_2.setText(_translate("MainWindow", "NamePosition"))
        self.label_4.setText(_translate("MainWindow", "BirthPosition"))
        self.Apply_btn.setText(_translate("MainWindow", "Apply"))
        self.Save_btn.setText(_translate("MainWindow", "SaveFile"))
        self.CharmGongik_btn.setText(_translate("MainWindow", "CharmGongik"))
        self.menuInfo.setTitle(_translate("MainWindow", "Info"))
        self.actionQT_License.setText(_translate("MainWindow", "QT License"))
        self.actionDeveloper_License.setText(_translate("MainWindow", "Developer License"))

    def ChamGongikView(self):
        dialog = QDialog()
        dialog.ui = CharmGongik()
        dialog.ui.setupUi(dialog)
        dialog.setAttribute(QtCore.Qt.WA_DeleteOnClose)
        dialog.exec_()
    
    def openSourceLicenseView(self):
        dialog = QDialog()
        dialog.ui = DialogLicense()
        dialog.ui.setupUi(dialog)
        dialog.setAttribute(QtCore.Qt.WA_DeleteOnClose)
        dialog.exec_()

    def doomsheartLicenseView(self):
        dialog = QDialog()
        dialog.ui = InfoLicense()
        dialog.ui.setupUi(dialog)
        dialog.setAttribute(QtCore.Qt.WA_DeleteOnClose)
        dialog.exec_()

    def exceptionDialog(self):
        dialog = QDialog()
        dialog.ui = ExceptionDialog()
        dialog.ui.setupUi(dialog)
        dialog.setAttribute(QtCore.Qt.WA_DeleteOnClose)
        dialog.exec_()

    def openNewMainFileDialog(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        files_, _ = QtWidgets.QFileDialog.getOpenFileNames(self.New_Main_File_btn, "FileBrowse", "",
                                                "Excel Files (*.xlsx);;", options=options)
        for file_ in files_:
            wb = openpyxl.load_workbook(file_)
            w = wb.active
            row_max_index = w.max_row
            column_max_index = w.max_column

            self.tableWidget_main.setRowCount(row_max_index)
            self.tableWidget_main.setColumnCount(column_max_index)

            self.tableWidget_Apply.setRowCount(row_max_index)
            self.tableWidget_Apply.setColumnCount(column_max_index)

            for row_index in range(row_max_index+1):
                for col_index in range(column_max_index+1):
                    self.tableWidget_main.setItem(row_index, col_index,
                                                  QTableWidgetItem(w.cell(row=row_index+1, column=col_index+1).value))
                    self.tableWidget_Apply.setItem(row_index, col_index,
                                                  QTableWidgetItem(w.cell(row=row_index+1, column=col_index+1).value))
         

        

        
    def openNewSubFileDialog(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        files_, _ = QtWidgets.QFileDialog.getOpenFileNames(self.New_Sub_File_btn, "FileBrowse", "",
                                                "Excel Files (*.xlsx);;", options=options)
        for file_ in files_:
            wb = openpyxl.load_workbook(file_)
            w = wb.active
            row_max_index = w.max_row
            column_max_index = w.max_column

            self.tableWidget_sub.setRowCount(row_max_index)
            self.tableWidget_sub.setColumnCount(column_max_index)

            for row_index in range(row_max_index+1):
                for col_index in range(column_max_index+1):
                    self.tableWidget_sub.setItem(row_index, col_index,
                                                 QTableWidgetItem(w.cell(row=row_index+1, column=col_index+1).value))

    def applyResult(self):
        main_name_pos = int(self.spinBox_Main_name_pos.text()) - 1
        main_birth_pos = int(self.spinBox_Main_birth_pos.text()) - 1
        sub_name_pos = int(self.spinBox_Sub_name_pos.text()) - 1
        sub_birth_pos = int(self.spinBox_Sub_birth_pos.text()) - 1
        target_info_pos = int(self.spinBox_Main_target_pos.text()) - 1

        main_info_list = []
        sub_info_list = []


        for i in range(self.tableWidget_main.rowCount()):
            main_info_list.append([])
        
        for i in range(self.tableWidget_sub.rowCount()):
            sub_info_list.append([])
        
        for row_index in range(self.tableWidget_Apply.rowCount()):
            for column_index in range(self.tableWidget_Apply.columnCount()):
                if self.tableWidget_Apply.item(row_index, column_index):
                    main_info_list[row_index].append(self.tableWidget_Apply.item(row_index, column_index).text().replace(" ", ""))
        
        # for row_index in range(self.tableWidget_main.rowCount()):
        #     for column_index in range(self.tableWidget_main.columnCount()):
        #         if self.tableWidget_main.item(row_index, column_index):
        #             main_info_list[row_index].append(self.tableWidget_main.item(row_index, column_index).text().replace(" ", ""))

        
        for row_index in range(self.tableWidget_sub.rowCount()):
            for column_index in range(self.tableWidget_sub.columnCount()):
                if self.tableWidget_sub.item(row_index, column_index):
                    sub_info_list[row_index].append(self.tableWidget_sub.item(row_index, column_index).text().replace(" ", ""))

        
        max_main_col = len(main_info_list) - 1
        max_main_row = 0
        try:
            max_main_row = len(main_info_list[0]) - 1
        except:
            max_main_row = -1
        

        max_sub_col = len(sub_info_list) - 1
        max_sub_row = -1
        try:
            max_sub_row = len(sub_info_list[0]) - 1
        except:
            max_sub_row = -1

        
        if (target_info_pos ==  max_main_row):
            for sub_info in sub_info_list:
                for main_info in main_info_list:
                    if (main_info[main_name_pos] == sub_info[sub_name_pos]) and (main_info[main_birth_pos] == sub_info[sub_birth_pos]):
                        main_info[target_info_pos] = main_info[target_info_pos] + str(sub_info[sub_name_pos]) + " : " + str(sub_info[sub_birth_pos])
                        continue
            

        elif(target_info_pos == max_main_row + 1):
            for main_info in main_info_list:
                main_info.append("")

            for sub_info in sub_info_list:
                for main_info in main_info_list:
                    if (main_info[main_name_pos] == sub_info[sub_name_pos]) and (main_info[main_birth_pos] == sub_info[sub_birth_pos]):
                        main_info[target_info_pos] = main_info[target_info_pos] + str(sub_info[sub_name_pos]) + " : " + str(sub_info[sub_birth_pos])
                        continue
        else:
            self.exceptionDialog()
            return


        row_max_index =  len(main_info_list)
        column_max_index = len(main_info_list[0])
        self.tableWidget_Apply.setRowCount(row_max_index)
        self.tableWidget_Apply.setColumnCount(column_max_index)

        for row_index in range(row_max_index):
            for col_index in range(column_max_index):
                self.tableWidget_Apply.setItem(row_index, col_index,
                                                QTableWidgetItem(main_info_list[row_index][col_index]))
    
    def saveResult(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        fileName, _ = QtWidgets.QFileDialog.getSaveFileName(self.Save_btn, "QFileDialog.getSaveFileName()", "",
                                                  "Excel File (*.xlsx)", options=options)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "인사변동 조사"
        result_list = []
        for row_index in range(self.tableWidget_Apply.rowCount()):
            result_list.append([])
        for row_index in range(self.tableWidget_Apply.rowCount()):
            for column_index in range(self.tableWidget_Apply.columnCount()):
                if self.tableWidget_Apply.item(row_index, column_index):
                    result_list[row_index].append(self.tableWidget_Apply.item(row_index, column_index).text().replace(" ", ""))
        for row in result_list:
            ws.append(row)
        # ws.append(result_list)
        wb.save(fileName)

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

